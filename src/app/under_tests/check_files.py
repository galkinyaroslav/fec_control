import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE_DIR = "/home/yaroslav/PycharmProjects/fec_control/src/data/runs"
OUTPUT_FILE = "cards_check.xlsx"


def check_pf(pf_dir, pf_label):
    """Проверка одной папки (0pF,10pF,20pF,40pF)"""
    result = {
        f"n_raw_{pf_label}": 0,
        f"n_gain_odd_{pf_label}": 0,
        f"n_gain_even_{pf_label}": 0,
        f"n_pll_{pf_label}": 0,
        f"n_pedest_{pf_label}": 0,
        f"{pf_label}_bad": False
    }

    if not os.path.isdir(pf_dir):
        result[f"{pf_label}_bad"] = True
        return result

    # raw
    raw_dir = os.path.join(pf_dir, "raw")
    if os.path.isdir(raw_dir):
        result[f"n_raw_{pf_label}"] = len(glob.glob(os.path.join(raw_dir, "*")))
    else:
        result[f"{pf_label}_bad"] = True

    # gain
    gain_dir = os.path.join(pf_dir, "gain")
    if os.path.isdir(gain_dir):
        even_files = glob.glob(os.path.join(gain_dir, "*-even-*"))
        odd_files = glob.glob(os.path.join(gain_dir, "*-odd-*"))
        result[f"n_gain_even_{pf_label}"] = len(even_files)
        result[f"n_gain_odd_{pf_label}"] = len(odd_files)

        if result[f"n_gain_even_{pf_label}"] == 0 or result[f"n_gain_odd_{pf_label}"] == 0:
            result[f"{pf_label}_bad"] = True
        if result[f"n_gain_even_{pf_label}"] % 32 != 0 or result[f"n_gain_odd_{pf_label}"] % 32 != 0:
            result[f"{pf_label}_bad"] = True
    else:
        result[f"{pf_label}_bad"] = True

    # pll
    pll_dir = os.path.join(pf_dir, "pll")
    if os.path.isdir(pll_dir):
        result[f"n_pll_{pf_label}"] = len(glob.glob(os.path.join(pll_dir, "*.pll")))
        if result[f"n_pll_{pf_label}"] == 0:
            result[f"{pf_label}_bad"] = True
    else:
        result[f"{pf_label}_bad"] = True

    # rms_pedestal
    ped_dir = os.path.join(pf_dir, "rms_pedestal")
    if os.path.isdir(ped_dir):
        result[f"n_pedest_{pf_label}"] = len(glob.glob(os.path.join(ped_dir, "*.txt")))
        if result[f"n_pedest_{pf_label}"] == 0:
            result[f"{pf_label}_bad"] = True
    else:
        result[f"{pf_label}_bad"] = True

    return result


def check_card(card_path, card_num):
    result = {"номер_карты": str(card_num), "bad": False}

    for pf in ["0pF", "10pF", "20pF", "40pF"]:
        pf_dir = os.path.join(card_path, pf)
        pf_result = check_pf(pf_dir, pf)
        result.update(pf_result)
        if pf_result[f"{pf}_bad"]:
            result["bad"] = True

    return result


def main():
    cards = []
    for entry in os.scandir(BASE_DIR):
        if entry.is_dir() and entry.name.isdigit():
            card_num = int(entry.name)
            cards.append(check_card(entry.path, card_num))

    df = pd.DataFrame(cards)

    # сохраняем в Excel
    df.to_excel(OUTPUT_FILE, index=False)

    # подсветка bad строк
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    bad_col = df.columns.get_loc("bad") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=bad_col).value:  # bad == True
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

    ws.delete_cols(bad_col)  # убираем служебный столбец
    wb.save(OUTPUT_FILE)
    print(f"Результат сохранён в {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
